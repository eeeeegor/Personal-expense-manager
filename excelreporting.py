from openpyxl import Workbook
from openpyxl.styles import Alignment
import banking


def full_excel_report(transactions, card_option, date):
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A2:D2')

    cells = 'ABCD'
    name = input('Choose filename (without .xlsx):')
    pos = 4
    year = int(date[:4])
    month = int(date[5:])
    card = -1

    ws['A3'] = 'Date'
    ws['B3'] = 'Type'
    ws['C3'] = 'Sum'
    ws['D4'] = 'Left'

    total_spent = 0
    total_received = 0

    print('year: ', year, 'month: ', month)

    if card_option != len(banking.CARD_MENU_INDEX.keys()) + 1:
        card = banking.CARD_MENU_INDEX[card_option][1]
    for temp in transactions:
        curr = temp
        if curr[-1][0] != year or curr[-1][1] != month:
            continue
        if card != -1:
            if curr[1] != card:
                continue

        date_to_write = '-'.join(map(str, curr[-1][:3])) + ' ' + ':'.join(map(str, curr[-1][3:]))
        type_of_transaction = 'Withdrawal' if curr[2] < 0 else 'Transfer'
        row = (date_to_write, type_of_transaction, str(abs(curr[2])), str(curr[3]))

        if type_of_transaction == 'Transfer':
            total_received += curr[2]
        else:
            total_spent += abs(curr[2])

        for i in range(4):
            place = cells[i] + str(pos)
            ws[place] = row[i]

        pos += 1

    ws['A' + str(pos + 2)] = 'Total received:'
    ws['A' + str(pos + 3)] = 'Total spent:'
    ws['B' + str(pos + 2)] = str(total_received)
    ws['B' + str(pos + 3)] = str(total_spent)

    wb.save(name + '.xlsx')




